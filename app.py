"""
TOSA – Transformer Oil Sample Analyzer  (Streamlit Web Version)
================================================================
Drop-in replacement for the Tkinter desktop app.
All core logic preserved: PDF extraction, ML predictions, scoring,
trend charts, dashboard, Need Attention table, and PDF report export.
"""

import os, re, math, json, io, warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
from scipy.interpolate import PchipInterpolator
import fitz                      # PyMuPDF
import joblib
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, Image as RLImage, PageBreak
)
import streamlit as st

warnings.filterwarnings("ignore")

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TOSA – Transformer Oil Analyzer",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).resolve().parent
ARTIFACT_DIR = BASE_DIR / "artifacts"
DB_PATH      = BASE_DIR / "Database.xlsx"
REPORTS_DIR  = BASE_DIR / "Reports"
REPORTS_DIR.mkdir(exist_ok=True)
ARTIFACT_DIR.mkdir(exist_ok=True)

# ── Download ML artifacts from Hugging Face (runs once, cached on disk) ────────
HF_REPO = "ICE-0027/TOSA-models"
HF_BASE = f"https://huggingface.co/{HF_REPO}/resolve/main"

ARTIFACT_FILES = [
    "rf_model_Oil_Quality.pkl",
    "rf_model_DGA1.pkl",
    "rf_model_DGA2.pkl",
    "label_encoder_Oil_Quality.pkl",
    "label_encoder_DGA1.pkl",
    "label_encoder_DGA2.pkl",
    "feature_meta.json",
]

@st.cache_resource(show_spinner=False)
def download_artifacts():
    import urllib.request
    missing = [f for f in ARTIFACT_FILES if not (ARTIFACT_DIR / f).is_file()]
    if missing:
        progress = st.progress(0, text="⬇️ Downloading ML models (first time only)…")
        for i, fname in enumerate(missing):
            url = f"{HF_BASE}/{fname}"
            dest = ARTIFACT_DIR / fname
            try:
                urllib.request.urlretrieve(url, dest)
            except Exception as e:
                st.error(f"Failed to download `{fname}` from Hugging Face.\n\n{e}\n\nCheck that the file exists at: {url}")
                st.stop()
            progress.progress((i + 1) / len(missing),
                               text=f"⬇️ Downloaded {i+1}/{len(missing)}: {fname}")
        progress.empty()

download_artifacts()

# ── Load ML artifacts ──────────────────────────────────────────────────────────
@st.cache_resource
def load_models():
    rf_oil  = joblib.load(ARTIFACT_DIR / "rf_model_Oil_Quality.pkl")
    rf_dga1 = joblib.load(ARTIFACT_DIR / "rf_model_DGA1.pkl")
    rf_dga2 = joblib.load(ARTIFACT_DIR / "rf_model_DGA2.pkl")
    le_oil  = joblib.load(ARTIFACT_DIR / "label_encoder_Oil_Quality.pkl")
    le_dga1 = joblib.load(ARTIFACT_DIR / "label_encoder_DGA1.pkl")
    le_dga2 = joblib.load(ARTIFACT_DIR / "label_encoder_DGA2.pkl")
    with open(ARTIFACT_DIR / "feature_meta.json", encoding="utf-8") as f:
        meta = json.load(f)
    return rf_oil, rf_dga1, rf_dga2, le_oil, le_dga1, le_dga2, meta

try:
    rf_oil, rf_dga1, rf_dga2, le_oil, le_dga1, le_dga2, _meta = load_models()
    FEATURES_OIL  = [str(c).strip() for c in _meta.get("Oil_Quality", [])]
    FEATURES_DGA1 = [str(c).strip() for c in _meta.get("DGA1", [])]
    FEATURES_DGA2 = [str(c).strip() for c in _meta.get("DGA2", [])]
    MODELS_OK = True
except Exception as _e:
    MODELS_OK = False
    st.error(f"⚠️  Could not load ML models.\n\n{_e}")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
RAW_2FAL_DROP_FOR_MAINT = 0.60
MAINT_NOTE_TEXT = "Oil filteration or replacment might occured prior this sample"
TREND_LINE_COLOR = "#1f77b4"

TREND_LIMITS = {
    "Power factor @25C/60Hz (%)": {"acceptable": (None,0.1),  "caution": (0.1,0.3),    "alarming": (0.3,None)},
    "IFT":                        {"acceptable": (28,None),   "caution": (22,28),       "alarming": (None,22)},
    "Breakdown voltage @2.5mm (kV)": {"acceptable":(30,None), "caution": (25,30),       "alarming": (None,25)},
    "Water Content @20 °C":       {"acceptable": (None,20),   "caution": (20,35),       "alarming": (35,None)},
    "Humidity in cellulose (%)":  {"acceptable": (None,2.0),  "caution": (2.0,3.0),     "alarming": (3.0,None)},
    "Acidity (mg KOH/g)":         {"acceptable": (None,0.06), "caution": (0.06,0.15),   "alarming": (0.15,None)},
    "Color ASTM":                 {"acceptable": (None,3),    "caution": (3,5),         "alarming": (5,None)},
    "DBDS":                       {"acceptable": (None,5),    "caution": (5,10),        "alarming": (10,None)},
    "Oil Quality Index":          {"acceptable": (2500,None), "caution": (1500,2500),   "alarming": (None,1500)},
    "Indirect DP-value":          {"acceptable": (500,None),  "caution": (350,500),     "alarming": (None,350)},
    "Hydrogen (H2)":              {"acceptable": (None,100),  "caution": (100,200),     "alarming": (200,None)},
    "Methane (CH4)":              {"acceptable": (None,90),   "caution": (90,150),      "alarming": (150,None)},
    "Ethane (C2H6)":              {"acceptable": (None,90),   "caution": (90,175),      "alarming": (175,None)},
    "Ethylene (C2H4)":            {"acceptable": (None,50),   "caution": (50,95),       "alarming": (95,None)},
    "Acetylene (C2H2)":           {"acceptable": (None,1),    "caution": (1,2),         "alarming": (2,None)},
    "TDCG":                       {"acceptable": (None,721),  "caution": (721,1920),    "alarming": (1920,None)},
    "Carbon Monoxide (CO)":       {"acceptable": (None,900),  "caution": (900,1100),    "alarming": (1100,None)},
    "Carbon Dioxide (CO2)":       {"acceptable": (None,10000),"caution": (10000,14000), "alarming": (14000,None)},
    "CO2/CO Ratio":               {"acceptable": (7.5,None),  "caution": (3,7.5),       "alarming": (None,3)},
}

PARAM_UNITS = {
    "Power factor @25C/60Hz (%)":"%" , "IFT":"mN/m",
    "Breakdown voltage @2.5mm (kV)":"kV", "Water Content @20 °C":"ppm",
    "Humidity in cellulose (%)":"%", "Acidity (mg KOH/g)":"mg KOH/g",
    "Oil Quality Index":"", "Color ASTM":"ASTM", "DBDS":"ppm",
    "Indirect DP-value":"", "Hydrogen (H2)":"ppm", "Methane (CH4)":"ppm",
    "Acetylene (C2H2)":"ppm", "Ethylene (C2H4)":"ppm", "Ethane (C2H6)":"ppm",
    "Carbon Monoxide (CO)":"ppm", "Carbon Dioxide (CO2)":"ppm",
    "TDCG":"ppm", "CO2/CO Ratio":"",
}
PARAM_LABELS = {
    "Power factor @25C/60Hz (%)":"PF (%)", "IFT":"IFT",
    "Breakdown voltage @2.5mm (kV)":"BDV", "Water Content @20 °C":"Moisture (oil)",
    "Humidity in cellulose (%)":"Moisture (paper)", "Acidity (mg KOH/g)":"Acidity",
    "Oil Quality Index":"Oil Quality Index", "Color ASTM":"Color", "DBDS":"DBDS",
    "Indirect DP-value":"DP-value", "Hydrogen (H2)":"H₂", "Methane (CH4)":"CH₄",
    "Acetylene (C2H2)":"C₂H₂", "Ethylene (C2H4)":"C₂H₄", "Ethane (C2H6)":"C₂H₆",
    "Carbon Monoxide (CO)":"CO", "Carbon Dioxide (CO2)":"CO₂",
    "TDCG":"TDCG", "CO2/CO Ratio":"CO₂/CO Ratio",
}

# ══════════════════════════════════════════════════════════════════════════════
#  HELPER UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _is_missing(v):
    if v is None: return True
    if isinstance(v, float) and math.isnan(v): return True
    return str(v).strip() in ("", "NA", "nan", "None")

def _to_float(v, default=0.0):
    try: return float(v) if not _is_missing(v) else default
    except: return default

def _to_int(v, default=0):
    try: return int(float(v)) if not _is_missing(v) else default
    except: return default

def _coerce_int(val):
    if _is_missing(val): return "NA"
    if isinstance(val, int): return val
    s = str(val).strip()
    if s.upper() == "LTD": return 0
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else "NA"

def _coerce_float(val):
    if _is_missing(val): return "NA"
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    if s.upper() == "LTD": return 0.0
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    return float(m.group(0)) if m else "NA"

def _as_excel(val):
    try:
        if val is None or (isinstance(val, float) and math.isnan(val)): return "NA"
    except: pass
    s = str(val).strip() if val is not None else ""
    return "NA" if s == "" else val

def clean_serial_number(sn):
    if sn is None: return "NA"
    cleaned = re.sub(r"[^A-Za-z0-9]", "", str(sn))
    return cleaned if cleaned else "NA"

def _norm_date_key(v):
    if v is None: return ""
    s = str(v).strip()
    if not s or s.upper() == "NA": return ""
    dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
    if pd.isna(dt): return s
    return dt.strftime("%d/%m/%Y")

def _norm_text_key(v):
    return str(v or "").strip().lower()

def _dedupe_pipe_text(s):
    parts = [p.strip() for p in str(s or "").split("|") if p.strip()]
    out = []
    for p in parts:
        if p not in out: out.append(p)
    return " | ".join(out)

def _clean_year_string(val):
    if val is None: return ""
    s = str(val).strip()
    m = re.search(r"\b(19|20)\d{2}\b", s)
    return m.group(0) if m else ""

def _parse_date_strict(s):
    s = (s or "").strip()
    m = re.match(r"^\s*(\d{1,2})[\/\-. ](\d{1,2})[\/\-. ](\d{4})\s*$", s)
    if not m: raise ValueError("Expected dd/mm/yyyy")
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    dt = pd.to_datetime(f"{d:02d}/{mo:02d}/{y:04d}", dayfirst=True, errors="coerce")
    if pd.isna(dt): raise ValueError("Invalid calendar date")
    return dt.strftime("%d/%m/%Y")

def parse_tension_hvlv(text):
    if not text: return "", None
    nums = re.findall(r"\d+(?:\.\d+)?", str(text))
    if not nums: return "", None
    return str(text).strip(), max(float(x) for x in nums)

def standardize_location(raw_loc, raw_sub=""):
    raw_loc = str(raw_loc or "").strip()
    raw_sub = str(raw_sub or "").strip()
    if not raw_loc or raw_loc.upper() == "NA": return raw_loc, raw_sub
    bracket = re.search(r'\(([^)]+)\)', raw_loc)
    if bracket:
        bt = bracket.group(1).strip()
        raw_loc = re.sub(r'\s*\([^)]+\)', '', raw_loc).strip()
        if bt:
            existing = raw_sub if (raw_sub and raw_sub.upper() != "NA") else ""
            raw_sub = (bt + " " + existing).strip() if existing else bt
    clean = re.sub(r'[^A-Za-z0-9 ]', ' ', raw_loc)
    clean = ' '.join(clean.split())
    upper = clean.upper()
    if 'SAF OFF' in upper: return "Safaniyah Offshore", raw_sub
    if 'SAF ON'  in upper: return "Safaniyah Onshore",  raw_sub
    if 'KPOD' in upper or 'KHPOD' in upper or 'KHUR' in upper: return "Khursaniyah", raw_sub
    if 'HARA' in upper: return "Haradh",      raw_sub
    if 'HAW'  in upper: return "Hawiyah",     raw_sub
    if 'UTH'  in upper: return "Uthmaniyah",  raw_sub
    if 'UDH'  in upper: return "Udhailiyah",  raw_sub
    if 'WEST' in upper or ' WR' in upper or upper.startswith('WR'): return "Western", raw_sub
    if 'RT'   in upper: return "Ras Tanura",  raw_sub
    if 'DH'   in upper: return "Dhahran",     raw_sub
    return clean if clean else raw_loc, raw_sub

# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_parameters(text):
    params = {
        "Serial Number:":"raw_string","SAP No.:":"raw_string","Location:":"raw_string",
        "Substation:":"raw_string","Tag Number:":"raw_string","Manufacture Name:":"raw_string",
        "Sampling Date:":"raw_string","Sampling From:":"raw_string",
        "Tension [HV/LV] (KV):":"raw_string","Manufacture Year:":"int",
        "Power Factor @25°C/60Hz":"float","Interfacial Tension (IFT)":"float",
        "Breakdown Voltage @2.5mm- AVG":"float","Humidity in Cellulose":"float",
        "Water Content @20 °C":"float","Total Acid Number":"float",
        "Oil Quality Index":"float","Color ASTM":"float",
        "Dibenzyl Disulfide (DBDS)":"float","2-Furfural (2FAL)":"float",
        "Indirect DP-value":"int","Hydrogen (H2)":"int","Methane (CH4)":"int",
        "Acetylene (C2H2)":"float","Ethylene (C2H4)":"float",
        "Ethane (C2H6)":"int","Carbon Monoxide (CO)":"int","Carbon Dioxide (CO2)":"int",
        "Report No. :":"raw_string",
    }
    extracted = {}
    for param, ptype in params.items():
        label = param.rstrip(":")
        m = re.search(re.escape(label) + r":?\s*([^\n]+)", text)
        if not m: continue
        value = m.group(1).strip()
        if ptype == "raw_string":
            extracted[param] = clean_serial_number(value) if param == "Serial Number:" else value
        elif ptype == "float":
            s = value.strip()
            if s.upper() == "LTD":
                extracted[param] = 0.0
            else:
                fm = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
                extracted[param] = float(fm.group(0)) if fm else "NA"
        elif ptype == "int":
            s = value.strip()
            if s.upper() == "LTD":
                extracted[param] = 0
            else:
                im = re.search(r"\d+", s)
                extracted[param] = int(im.group(0)) if im else "NA"
        if param == "Manufacture Year:":
            cy = _clean_year_string(value)
            extracted[param] = cy if cy else "NA"

    for k in ["Serial Number:","SAP No.:","Location:","Substation:","Tag Number:",
               "Sampling From:","Manufacture Name:","Tension [HV/LV] (KV):"]:
        if not str(extracted.get(k,"")).strip():
            extracted[k] = "NA"

    sd = str(extracted.get("Sampling Date:","") or "").strip()
    extracted["Sampling Date:"] = sd if sd else "NA"

    if "Sample No" not in extracted:
        rep = extracted.get("Report No. :","")
        if str(rep).strip(): extracted["Sample No"] = rep
    extracted.pop("Report No. :", None)
    return extracted

def extract_from_excel_row(row):
    ed = {}
    ed["Serial Number:"]         = clean_serial_number(str(row.get("SN","") or ""))
    ed["SAP No.:"]               = _coerce_raw_na(row.get("SAP No.:"))
    ed["Sample No"]              = _coerce_raw_na(row.get("Sample No"))
    ed["Location:"]              = _coerce_raw_na(row.get("Location:"))
    ed["Substation:"]            = _coerce_raw_na(row.get("Substation:"))
    ed["Tag Number:"]            = _coerce_raw_na(row.get("Tag Number:"))
    ed["Sampling From:"]         = _coerce_raw_na(row.get("Sampling From:"))
    ed["Manufacture Name:"]      = _coerce_raw_na(row.get("Manufacture Name:"))
    ed["Tension [HV/LV] (KV):"] = _coerce_raw_na(row.get("Tension [HV/LV] (KV):"))
    _sd = _coerce_date_ddmmyyyy(row.get("Sampling Date:"))
    ed["Sampling Date:"] = _sd if _sd else "NA"
    raw_my = row.get("Manufacture Year:")
    cy = _clean_year_string(raw_my)
    ed["Manufacture Year:"] = cy if cy else "NA"
    ed["Indirect DP-value"]          = _coerce_int(row.get("Indirect DP-value"))
    ed["Hydrogen (H2)"]              = _coerce_int(row.get("Hydrogen (H2)"))
    ed["Methane (CH4)"]              = _coerce_int(row.get("Methane (CH4)"))
    ed["Ethane (C2H6)"]              = _coerce_int(row.get("Ethane (C2H6)"))
    ed["Carbon Monoxide (CO)"]       = _coerce_int(row.get("Carbon Monoxide (CO)"))
    ed["Carbon Dioxide (CO2)"]       = _coerce_int(row.get("Carbon Dioxide (CO2)"))
    ed["Oil Quality Index"]          = _coerce_float(row.get("Oil Quality Index"))
    ed["Power Factor @25°C/60Hz"]    = _coerce_float(row.get("Power Factor @25°C/60Hz"))
    ed["Interfacial Tension (IFT)"]  = _coerce_float(row.get("Interfacial Tension (IFT)"))
    ed["Breakdown Voltage @2.5mm- AVG"] = _coerce_float(row.get("Breakdown Voltage @2.5mm- AVG"))
    ed["Humidity in Cellulose"]      = _coerce_float(row.get("Humidity in Cellulose"))
    ed["Water Content @20 °C"]       = _coerce_float(row.get("Water Content @20 °C"))
    ed["Total Acid Number"]          = _coerce_float(row.get("Total Acid Number"))
    ed["Color ASTM"]                 = _coerce_float(row.get("Color ASTM"))
    ed["Dibenzyl Disulfide (DBDS)"]  = _coerce_float(row.get("Dibenzyl Disulfide (DBDS)"))
    ed["2-Furfural (2FAL)"]          = _coerce_float(row.get("2-Furfural (2FAL)"))
    ed["Acetylene (C2H2)"]           = _coerce_float(row.get("Acetylene (C2H2)"))
    ed["Ethylene (C2H4)"]            = _coerce_float(row.get("Ethylene (C2H4)"))
    loc, sub = standardize_location(ed.get("Location:",""), ed.get("Substation:",""))
    ed["Location:"] = loc; ed["Substation:"] = sub
    return ed

def _coerce_raw_na(val):
    if val is None: return "NA"
    s = str(val).strip()
    return s if s else "NA"

def _coerce_date_ddmmyyyy(v):
    if v is None: return ""
    s = str(v).strip()
    if not s or s.upper() == "NA": return ""
    dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
    if pd.isna(dt): return ""
    return dt.strftime("%d/%m/%Y")

def _normalize_sampling_date(extracted_data):
    sd = str(extracted_data.get("Sampling Date:","") or "").strip()
    if not sd or sd.upper() == "NA":
        extracted_data["Sampling Date:"] = "NA"
        return
    dt = pd.to_datetime(sd, dayfirst=True, errors="coerce")
    if not pd.isna(dt):
        extracted_data["Sampling Date:"] = dt.strftime("%d/%m/%Y")

# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE (Excel)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=5)
def _load_db_cached():
    if not DB_PATH.is_file(): return None
    try:
        df = pd.read_excel(DB_PATH)
        df.columns = df.columns.astype(str).str.strip()
        if "Sampling Date" in df.columns:
            df["Sampling Date"] = pd.to_datetime(df["Sampling Date"], dayfirst=True, errors="coerce")
        if "SN" in df.columns:
            df["SN"] = df["SN"].astype(str).apply(clean_serial_number)
        return df
    except: return None

def load_db():
    _load_db_cached.clear()
    return _load_db_cached()

def _latest_row_by_key(df, key_col, key_val, curr_date_str):
    if df is None or key_col not in df.columns or _is_missing(key_val): return None
    key_val = str(key_val).strip()
    if key_col == "SN":
        kvc = clean_serial_number(key_val)
        subset = df[df["SN"].astype(str).apply(clean_serial_number) == kvc]
    else:
        subset = df[df[key_col].astype(str).str.strip() == key_val]
    if subset.empty: return None
    if "Sampling Date" in subset.columns:
        curr_date = pd.to_datetime(curr_date_str, dayfirst=True, errors="coerce")
        subset = subset.copy()
        if not pd.isna(curr_date):
            older = subset[pd.to_datetime(subset["Sampling Date"], errors="coerce") < curr_date]
            if not older.empty: subset = older
        subset = subset.sort_values("Sampling Date")
    return subset.tail(1).iloc[0]

_INV_MAP = {
    "Location:":"Location","Substation:":"Substation","Tag Number:":"Tag Number",
    "Manufacture Name:":"Manufacture Name","Sampling From:":"Sampling From",
    "Tension [HV/LV] (KV):":"Tension [HV/LV] (KV)",
    "Power Factor @25°C/60Hz":"Power factor @25C/60Hz (%)",
    "Interfacial Tension (IFT)":"IFT",
    "Breakdown Voltage @2.5mm- AVG":"Breakdown voltage @2.5mm (kV)",
    "Humidity in Cellulose":"Humidity in cellulose (%)",
    "Water Content @20 °C":"Water Content @20 °C",
    "Total Acid Number":"Acidity (mg KOH/g)",
    "Oil Quality Index":"Oil Quality Index",
    "Color ASTM":"Color ASTM","Dibenzyl Disulfide (DBDS)":"DBDS",
    "2-Furfural (2FAL)":"2-Furfural (2FAL)",
    "Indirect DP-value":"Indirect DP-value",
    "Hydrogen (H2)":"Hydrogen (H2)","Methane (CH4)":"Methane (CH4)",
    "Acetylene (C2H2)":"Acetylene (C2H2)","Ethylene (C2H4)":"Ethylene (C2H4)",
    "Ethane (C2H6)":"Ethane (C2H6)","Carbon Monoxide (CO)":"Carbon Monoxide (CO)",
    "Carbon Dioxide (CO2)":"Carbon Dioxide (CO2)",
}

def backfill_metadata(extracted_data, curr_date_str):
    df = _load_db_cached()
    sn = str(extracted_data.get("Serial Number:","") or "").strip()
    sap = str(extracted_data.get("SAP No.:","") or "").strip()
    if _is_missing(sn):
        if not _is_missing(sap) and df is not None and "SAP No" in df.columns:
            row = _latest_row_by_key(df, "SAP No", sap, curr_date_str)
            if row is not None and "SN" in row and not _is_missing(row["SN"]):
                extracted_data["Serial Number:"] = clean_serial_number(str(row["SN"]).strip())
            else: extracted_data["Serial Number:"] = "NA"
        else: extracted_data["Serial Number:"] = "NA"
    sn = str(extracted_data.get("Serial Number:","") or "").strip()
    if _is_missing(sap):
        if not _is_missing(sn) and df is not None and "SN" in df.columns:
            row = _latest_row_by_key(df, "SN", sn, curr_date_str)
            if row is not None and "SAP No" in row and not _is_missing(row["SAP No"]):
                extracted_data["SAP No.:"] = str(row["SAP No"]).strip()
            else: extracted_data["SAP No.:"] = "NA"
        else: extracted_data["SAP No.:"] = "NA"
    sap = str(extracted_data.get("SAP No.:","") or "").strip()
    for key in ["Location:","Substation:","Tag Number:","Manufacture Name:","Sampling From:","Tension [HV/LV] (KV):"]:
        if not _is_missing(extracted_data.get(key,"")): continue
        db_col = _INV_MAP.get(key)
        if not db_col or df is None or db_col not in df.columns:
            extracted_data[key] = "NA"; continue
        val = None
        if not _is_missing(sap) and "SAP No" in df.columns:
            row = _latest_row_by_key(df, "SAP No", sap, curr_date_str)
            if row is not None and db_col in row and not _is_missing(row[db_col]):
                val = str(row[db_col]).strip()
        if val is None and not _is_missing(sn) and "SN" in df.columns:
            row = _latest_row_by_key(df, "SN", sn, curr_date_str)
            if row is not None and db_col in row and not _is_missing(row[db_col]):
                val = str(row[db_col]).strip()
        extracted_data[key] = val if val else "NA"

_FLOAT_PARAMS = {"Power Factor @25°C/60Hz","Interfacial Tension (IFT)","Breakdown Voltage @2.5mm- AVG",
                 "Humidity in Cellulose","Water Content @20 °C","Total Acid Number","Oil Quality Index",
                 "Color ASTM","Dibenzyl Disulfide (DBDS)","2-Furfural (2FAL)","Acetylene (C2H2)","Ethylene (C2H4)"}
_INT_PARAMS = {"Indirect DP-value","Hydrogen (H2)","Methane (CH4)","Ethane (C2H6)",
               "Carbon Monoxide (CO)","Carbon Dioxide (CO2)"}

def backfill_numeric(param, sap, sn, curr_date_str):
    df = _load_db_cached()
    if df is None: return None
    db_col = _INV_MAP.get(param, param)
    if db_col not in df.columns: return None
    for key_col, key_val in [("SAP No", sap), ("SN", sn)]:
        if _is_missing(key_val) or key_col not in df.columns: continue
        row = _latest_row_by_key(df, key_col, str(key_val), curr_date_str)
        if row is None or db_col not in row or _is_missing(row[db_col]): continue
        v = row[db_col]
        if param in _INT_PARAMS:
            c = _coerce_int(v); return None if c == "NA" else c
        elif param in _FLOAT_PARAMS:
            c = _coerce_float(v); return None if c == "NA" else c
        return v
    return None

REQUIRED_PARAMS = [
    "Power Factor @25°C/60Hz","Interfacial Tension (IFT)","Breakdown Voltage @2.5mm- AVG",
    "Humidity in Cellulose","Water Content @20 °C","Total Acid Number","Oil Quality Index",
    "Color ASTM","Dibenzyl Disulfide (DBDS)","2-Furfural (2FAL)","Indirect DP-value",
    "Hydrogen (H2)","Methane (CH4)","Acetylene (C2H2)","Ethylene (C2H4)","Ethane (C2H6)",
    "Carbon Monoxide (CO)","Carbon Dioxide (CO2)",
]
REQ_TYPES = {
    "Power Factor @25°C/60Hz":"float","Interfacial Tension (IFT)":"float",
    "Breakdown Voltage @2.5mm- AVG":"float","Humidity in Cellulose":"float",
    "Water Content @20 °C":"float","Total Acid Number":"float","Oil Quality Index":"float",
    "Color ASTM":"float","Dibenzyl Disulfide (DBDS)":"float","2-Furfural (2FAL)":"float",
    "Indirect DP-value":"int","Hydrogen (H2)":"int","Methane (CH4)":"int",
    "Acetylene (C2H2)":"float","Ethylene (C2H4)":"float","Ethane (C2H6)":"int",
    "Carbon Monoxide (CO)":"int","Carbon Dioxide (CO2)":"int",
}

def _needs_prompt(val, ptype):
    if _is_missing(val): return True
    if ptype in ("int","float"):
        if isinstance(val,(int,float)): return False
        s = str(val).strip()
        return not re.search(r"\d", s)
    return False

# ══════════════════════════════════════════════════════════════════════════════
#  SCORING
# ══════════════════════════════════════════════════════════════════════════════

def score_v2(row):
    BDV = _to_float(row.get("Breakdown voltage @2.5mm (kV)"))
    IFT = _to_float(row.get("IFT"))
    Acid = _to_float(row.get("Acidity (mg KOH/g)"))
    PF = _to_float(row.get("Power factor @25C/60Hz (%)"))
    Moisture_in_oil = _to_float(row.get("Water Content @20 °C"))
    Moisture_in_paper = _to_float(row.get("Humidity in cellulose (%)"))
    DBDS = _to_float(row.get("DBDS"))
    Color = _to_float(row.get("Color ASTM"))
    Hydrogen = _to_float(row.get("Hydrogen (H2)"))
    Methane = _to_float(row.get("Methane (CH4)"))
    Accetylene = _to_float(row.get("Acetylene (C2H2)"))
    Ethylene = _to_float(row.get("Ethylene (C2H4)"))
    Ethane = _to_float(row.get("Ethane (C2H6)"))
    CO = _to_float(row.get("Carbon Monoxide (CO)"))
    TDCG = _to_float(row.get("TDCG"))
    CO2 = _to_float(row.get("Carbon Dioxide (CO2)"))
    CO2_CO_ratio = _to_float(row.get("CO2/CO Ratio"))
    DP = _to_int(row.get("Indirect DP-value"))

    BDV_score = (0 if BDV<25 else 1 if BDV<27.5 else 2 if BDV<29.5 else 3)
    IFT_score = (0 if IFT<28 else 1 if IFT<29.5 else 2 if IFT<32 else 3)
    Acidity_score = (3 if Acid<0.06 else 2 if Acid<0.085 else 1 if Acid<0.15 else 0)
    PF_score = (3 if PF<0.1 else 2 if PF<0.21 else 1 if PF<0.31 else 0)
    Water_content_score = (3 if Moisture_in_oil<10 else 2 if Moisture_in_oil<20 else 1 if Moisture_in_oil<26 else 0)
    Moisture_in_paper_score = (3 if Moisture_in_paper<1.26 else 2 if Moisture_in_paper<2.01 else 1 if Moisture_in_paper<2.51 else 0)
    Corrosive = (3 if DBDS<10 else 0)
    Color_score = (3 if Color<3.51 else 0)
    Quality_Index = _to_float(row.get("Oil Quality Index"), 0.0)
    Quality_Index_score = (0 if Quality_Index<1501 else 3)
    Hydrogen_score = (3 if Hydrogen<101 else 2 if Hydrogen<701 else 1 if Hydrogen<1801 else 0)
    Methane_score = (3 if Methane<121 else 2 if Methane<401 else 1 if Methane<1001 else 0)
    Accetylene_score = (3 if Accetylene<2 else 2 if Accetylene<10 else 1 if Accetylene<36 else 0)
    Ethylene_score = (3 if Ethylene<51 else 2 if Ethylene<101 else 1 if Ethylene<201 else 0)
    Ethane_score = (3 if Ethane<66 else 2 if Ethane<101 else 1 if Ethane<151 else 0)
    CO_score = (3 if CO<351 else 2 if CO<571 else 1 if CO<1401 else 0)
    TDCG_score = (3 if TDCG<721 else 2 if TDCG<1921 else 1 if TDCG<4631 else 0)
    CO2_score = (3 if CO2<2501 else 2 if CO2<4001 else 1 if CO2<10001 else 0)
    CO2_CO_ratio_score = (0 if CO2_CO_ratio<3 else 1 if CO2_CO_ratio<5 else 2 if CO2_CO_ratio<7.5 else 3)

    IFT_Acidity_zeros = sum(x==0 for x in [IFT_score,Acidity_score])
    IFT_Acidity_ones  = sum(x==1 for x in [IFT_score,Acidity_score])
    Moisture_in_paper_zeros = sum(x==0 for x in [Moisture_in_paper_score])
    Moisture_in_paper_ones  = sum(x==1 for x in [Moisture_in_paper_score])
    Hydrogen_Methane_zeros = sum(x==0 for x in [Hydrogen_score,Methane_score])
    Hydrogen_Methane_ones  = sum(x==1 for x in [Hydrogen_score,Methane_score])
    DGA_zeros = sum(x==0 for x in [Ethylene_score,Ethane_score,CO_score,TDCG_score])
    DGA_ones  = sum(x==1 for x in [Ethylene_score,Ethane_score,CO_score,TDCG_score])

    formula = (1.5**(Hydrogen_Methane_ones+DGA_ones))*(2**(Hydrogen_Methane_zeros+DGA_zeros))
    DGA_sum = ((Hydrogen_score*0.2)+(Methane_score*0.017)+(Accetylene_score*0.5)+
               (Ethylene_score*0.0166)+(Ethane_score*0.0166)+(CO_score*0.0166)+
               (TDCG_score*0.2)+(CO2_score*0.0166)+(CO2_CO_ratio_score*0.0166))

    if Accetylene_score>1:   DGA_score = DGA_sum/formula if formula else 0
    elif Accetylene_score==1: DGA_score = DGA_sum/(2*formula) if formula else 0
    else:                     DGA_score = DGA_sum/(3*formula) if formula else 0

    Oil_quality_score = ((BDV_score*0.1)+(IFT_score*0.155)+(Acidity_score*0.2)+(PF_score*0.1)+
        (Water_content_score*0.07)+(Moisture_in_paper_score*0.08)+(Corrosive*0.2)+
        (Color_score*0.045)+(Quality_Index_score*0.05)
    ) / ((2**IFT_Acidity_zeros)*(1.5**IFT_Acidity_ones)*(1.5**Moisture_in_paper_zeros)*(1.25**Moisture_in_paper_ones))

    Oil_quality_score *= 0.91
    DGA_score *= 0.61
    DP_score = (0 if DP<360 else 1 if DP<600 else 2 if DP<800 else 3)*1.5
    return round(Oil_quality_score+DGA_score+DP_score+0.93, 2)

# ══════════════════════════════════════════════════════════════════════════════
#  RECOMPUTE HISTORY (2FAL correction, ML re-run)
# ══════════════════════════════════════════════════════════════════════════════

def recompute_transformer_history(sn):
    if not DB_PATH.is_file(): return
    df = pd.read_excel(DB_PATH)
    df.columns = df.columns.astype(str).str.strip()
    if "SN" not in df.columns: return
    df["SN"] = df["SN"].astype(str).apply(clean_serial_number)
    if "Sampling Date" in df.columns:
        df["Sampling Date"] = pd.to_datetime(df["Sampling Date"], dayfirst=True, errors="coerce")

    mask = df["SN"] == sn
    sub = df[mask].copy()
    if "Sampling From" in sub.columns:
        sub = sub[sub["Sampling From"].astype(str).str.lower().str.contains("bottom", na=False)]
    if sub.empty: return

    sub = sub.sort_values("Sampling Date")
    sub["_dt"] = pd.to_datetime(sub["Sampling Date"], errors="coerce")
    sub = sub.sort_values("_dt")

    prev_raw = None; peak_actual = 0.0; latest_actual_before_replacement = 0.0

    for i in sub.index:
        curr_raw_raw = sub.at[i, "Raw 2-Furfural (2FAL)"] if "Raw 2-Furfural (2FAL)" in sub.columns else None
        if _is_missing(curr_raw_raw):
            curr_raw_raw = sub.at[i, "2-Furfural (2FAL)"] if "2-Furfural (2FAL)" in sub.columns else 0
        try: curr_raw = float(curr_raw_raw) if not _is_missing(curr_raw_raw) else 0.0
        except: curr_raw = 0.0

        raw_dp_reported = sub.at[i, "Raw Indirect DP"] if "Raw Indirect DP" in sub.columns else None
        if _is_missing(raw_dp_reported):
            raw_dp_reported = sub.at[i, "Indirect DP-value"] if "Indirect DP-value" in sub.columns else None
        sub.at[i, "Raw Indirect DP"] = raw_dp_reported

        maint_reset = False
        if prev_raw is not None and prev_raw > 1:
            drop_frac = (prev_raw - curr_raw)/prev_raw if prev_raw > 0 else 0.0
            maint_reset = drop_frac >= RAW_2FAL_DROP_FOR_MAINT
        if maint_reset: latest_actual_before_replacement = peak_actual

        actual = curr_raw + latest_actual_before_replacement
        if actual < peak_actual: actual = peak_actual
        else: peak_actual = actual
        sub.at[i, "2-Furfural (2FAL)"] = actual
        sub.at[i, "Maintenance Note"] = MAINT_NOTE_TEXT if maint_reset else ""

        if actual > 0:
            dp_corr = int(round((1.5655 - math.log10(actual))/0.0035))
        else: dp_corr = 0
        sub.at[i, "Indirect DP-value"] = dp_corr

        # Voltage
        tension_str = sub.at[i, "Tension [HV/LV] (KV)"] if "Tension [HV/LV] (KV)" in sub.columns else ""
        _, voltage_level = parse_tension_hvlv(str(tension_str))
        if voltage_level is None: voltage_level = 0

        # Age
        age_val = _to_int(sub.at[i, "Age"] if "Age" in sub.columns else 0)

        # ML input
        def _get(col): return sub.at[i, col] if col in sub.columns else 0

        ml_oil = {
            "Voltage level (kV)":[voltage_level], "Age":[age_val],
            "Power factor @25C/60Hz (%)":[_get("Power factor @25C/60Hz (%)")],
            "IFT":[_get("IFT")], "Breakdown voltage @2.5mm (kV)":[_get("Breakdown voltage @2.5mm (kV)")],
            "Humidity in cellulose (%)":[_get("Humidity in cellulose (%)")],
            "Water Content @20 °C":[_get("Water Content @20 °C")],
            "Acidity (mg KOH/g)":[_get("Acidity (mg KOH/g)")],
            "Oil Quality Index":[_get("Oil Quality Index")],
            "Color ASTM":[_get("Color ASTM")], "DBDS":[_get("DBDS")],
            "Indirect DP-value":[dp_corr],
        }
        ml_dga = {
            "Hydrogen (H2)":[_get("Hydrogen (H2)")], "Methane (CH4)":[_get("Methane (CH4)")],
            "Acetylene (C2H2)":[_get("Acetylene (C2H2)")], "Ethylene (C2H4)":[_get("Ethylene (C2H4)")],
            "Ethane (C2H6)":[_get("Ethane (C2H6)")], "Carbon Monoxide (CO)":[_get("Carbon Monoxide (CO)")],
            "Carbon Dioxide (CO2)":[_get("Carbon Dioxide (CO2)")], "TDCG":[_get("TDCG")],
            "CO2/CO Ratio":[_get("CO2/CO Ratio")], "Indirect DP-value":[dp_corr],
        }
        try:
            df_oil = pd.DataFrame(ml_oil).reindex(columns=FEATURES_OIL).fillna(0)
            sub.at[i,"Oil Quality Recommendations"] = str(le_oil.inverse_transform(rf_oil.predict(df_oil))[0])
        except: pass
        try:
            df_d1 = pd.DataFrame(ml_dga).reindex(columns=FEATURES_DGA1).fillna(0)
            sub.at[i,"Group1 DGA Recommendations"] = str(le_dga1.inverse_transform(rf_dga1.predict(df_d1))[0])
        except: pass
        try:
            df_d2 = pd.DataFrame(ml_dga).reindex(columns=FEATURES_DGA2).fillna(0)
            sub.at[i,"Group2 DGA Recommendations"] = str(le_dga2.inverse_transform(rf_dga2.predict(df_d2))[0])
        except: pass

        score_row = {k: sub.at[i, k] if k in sub.columns else 0 for k in [
            "Breakdown voltage @2.5mm (kV)","IFT","Acidity (mg KOH/g)",
            "Power factor @25C/60Hz (%)","Water Content @20 °C","Humidity in cellulose (%)",
            "DBDS","Color ASTM","Oil Quality Index","Hydrogen (H2)","Methane (CH4)",
            "Acetylene (C2H2)","Ethylene (C2H4)","Ethane (C2H6)",
            "Carbon Monoxide (CO)","TDCG","Carbon Dioxide (CO2)","CO2/CO Ratio","Indirect DP-value"]}
        s = score_v2(score_row)
        sub.at[i,"Score"] = s
        sub.at[i,"Condition"] = ("Good" if s>=6.67 else "Fair" if s>=3.34 else "Poor")
        sub.at[i,"Escalated to TCA L2"] = "Yes" if s<6.67 else "No"
        prev_raw = curr_raw

    for col in sub.columns:
        df.loc[sub.index, col] = sub[col]
    df.drop(columns=["_dt"], errors="ignore", inplace=True)
    df.to_excel(DB_PATH, index=False)

# ══════════════════════════════════════════════════════════════════════════════
#  SAVE TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def save_to_excel(extracted_data, oil_rec, dga1_rec, dga2_rec, maint_note=""):
    sn_key  = clean_serial_number(str(extracted_data.get("Serial Number:","")).strip())
    sd_key  = _norm_date_key(extracted_data.get("Sampling Date:",""))
    sf_key  = _norm_text_key(extracted_data.get("Sampling From:",""))
    sno_key = _norm_text_key(extracted_data.get("Sample No",""))

    if DB_PATH.is_file():
        try:
            df_db = pd.read_excel(DB_PATH)
            df_db.columns = df_db.columns.astype(str).str.strip()
            for col,fn in [("SN", lambda v: clean_serial_number(str(v))),
                           ("Sampling Date", _norm_date_key),
                           ("Sampling From", _norm_text_key),
                           ("Sample No", _norm_text_key)]:
                if col in df_db.columns: df_db[col] = df_db[col].apply(fn)
                else: df_db[col] = ""
            dup1 = ((df_db["SN"]==sn_key) & (df_db["Sampling Date"]==sd_key) & (df_db["Sampling From"]==sf_key))
            dup2 = ((df_db["SN"]==sn_key) & (df_db["Sample No"]==sno_key)) if sno_key else False
            if dup1.any() or (hasattr(dup2,"any") and dup2.any()): return False
        except: pass

    voltage_text = extracted_data.get("Tension [HV/LV] (KV):","")
    m = re.search(r"(\d+(?:\.\d+)?)", str(voltage_text))
    voltage_level = float(m.group(1)) if m else None

    co2 = extracted_data.get("Carbon Dioxide (CO2)")
    co  = extracted_data.get("Carbon Monoxide (CO)")
    h2  = extracted_data.get("Hydrogen (H2)")
    ch4 = extracted_data.get("Methane (CH4)")
    c2h6= extracted_data.get("Ethane (C2H6)")
    c2h2= extracted_data.get("Acetylene (C2H2)")
    c2h4= extracted_data.get("Ethylene (C2H4)")
    vals = [h2,ch4,c2h6,c2h2,c2h4,co]
    TDCG = sum(_to_float(v,0.0) for v in vals)
    co2_co_ratio = round(co2/co,2) if (isinstance(co2,(int,float)) and isinstance(co,(int,float)) and co!=0) else None
    dp_now = extracted_data.get("Indirect DP-value")
    try: rem_life = int(round(float(dp_now)*30.0/1000.0-1.0)) if dp_now is not None else None
    except: rem_life = None

    age = extracted_data.get("Age","NA")
    score = extracted_data.get("Score",0)
    condition = ("Good" if isinstance(score,float) and score>=6.67 else
                 "Fair" if isinstance(score,float) and score>=3.34 else
                 "Poor" if isinstance(score,float) else "")
    escalated = "Yes" if isinstance(score,float) and score<6.67 else "No"

    columns = [
        "SN","SAP No","Sample No","Location","Substation","Tag Number","Manufacture Name",
        "Sampling Date","Sampling From","Tension [HV/LV] (KV)","Age",
        "Power factor @25C/60Hz (%)","IFT","Breakdown voltage @2.5mm (kV)",
        "Humidity in cellulose (%)","Water Content @20 °C","Acidity (mg KOH/g)",
        "Oil Quality Index","Color ASTM","DBDS","2-Furfural (2FAL)","Indirect DP-value","Remaining Lifetime",
        "Hydrogen (H2)","Methane (CH4)","Acetylene (C2H2)","Ethylene (C2H4)","Ethane (C2H6)",
        "Carbon Monoxide (CO)","Carbon Dioxide (CO2)","TDCG","CO2/CO Ratio",
        "Oil Quality Recommendations","Group1 DGA Recommendations","Group2 DGA Recommendations",
        "Maintenance Note","Score","Condition","Escalated to TCA L2",
        "Raw 2-Furfural (2FAL)","Raw Indirect DP","Raw Score"
    ]
    row = [
        extracted_data.get("Serial Number:",""), extracted_data.get("SAP No.:",""),
        extracted_data.get("Sample No",""), extracted_data.get("Location:",""),
        extracted_data.get("Substation:",""), extracted_data.get("Tag Number:",""),
        extracted_data.get("Manufacture Name:",""), extracted_data.get("Sampling Date:",""),
        extracted_data.get("Sampling From:",""), extracted_data.get("Tension [HV/LV] (KV):",""),
        age, extracted_data.get("Power Factor @25°C/60Hz"),
        extracted_data.get("Interfacial Tension (IFT)"),
        extracted_data.get("Breakdown Voltage @2.5mm- AVG"),
        extracted_data.get("Humidity in Cellulose"),
        extracted_data.get("Water Content @20 °C"),
        extracted_data.get("Total Acid Number"),
        extracted_data.get("Oil Quality Index"),
        extracted_data.get("Color ASTM"),
        extracted_data.get("Dibenzyl Disulfide (DBDS)"),
        extracted_data.get("2-Furfural (2FAL)"),
        extracted_data.get("Indirect DP-value"), rem_life,
        h2, ch4, c2h2, c2h4, c2h6, co, co2, TDCG, co2_co_ratio,
        oil_rec, dga1_rec, dga2_rec, maint_note,
        score, condition, escalated,
        extracted_data.get("2-Furfural (2FAL)"),
        extracted_data.get("Indirect DP-value"), score
    ]
    wb = Workbook() if not DB_PATH.is_file() else load_workbook(DB_PATH)
    ws = wb.active
    if not DB_PATH.is_file() or ws.max_row == 1:
        ws.append(columns)
    ws.append([_as_excel(v) for v in row])
    date_col = columns.index("Sampling Date")+1
    date_str = str(extracted_data.get("Sampling Date:","")).strip()
    try:
        dt = datetime.strptime(date_str, "%d/%m/%Y")
        c = ws.cell(row=ws.max_row, column=date_col)
        c.value = dt; c.number_format = "DD/MM/YYYY"
    except: pass
    wb.save(DB_PATH)
    return True

# ══════════════════════════════════════════════════════════════════════════════
#  TREND CHARTS
# ══════════════════════════════════════════════════════════════════════════════

def _catmull_rom_chain(x, y, samples_per_seg=25):
    x = np.asarray(x, dtype=float); y = np.asarray(y, dtype=float)
    if len(x)<2: return x, y
    _, uid = np.unique(x, return_index=True)
    x = x[uid]; y = y[uid]
    if len(x)<2: return x, y
    xs = np.linspace(x[0], x[-1], (len(x)-1)*samples_per_seg+1)
    return xs, PchipInterpolator(x,y)(xs)

def _add_trend_bands(ax, param_name, ymin_plot, ymax_plot):
    limits = TREND_LIMITS.get(param_name)
    if not limits: return
    for key, color in [("acceptable","#4CAF50"),("caution","#FFA726"),("alarming","#EF5350")]:
        low, high = limits.get(key,(None,None))
        band_low  = ymin_plot if low  is None else low
        band_high = ymax_plot if high is None else high
        band_low  = max(band_low, ymin_plot)
        band_high = min(band_high, ymax_plot)
        if band_high<=band_low: continue
        ax.axhspan(band_low, band_high, facecolor=color, alpha=0.45, zorder=0)

def generate_trend_graphs(serial_number, db_df=None):
    if db_df is None:
        if not DB_PATH.is_file(): return []
        db_df = pd.read_excel(DB_PATH)
        db_df.columns = db_df.columns.astype(str).str.strip()
    df = db_df.copy()
    if "Sampling From" in df.columns:
        df = df[df["Sampling From"].astype(str).str.lower().str.contains("bottom", na=False)]
    df["Sampling Date"] = pd.to_datetime(df["Sampling Date"], dayfirst=True, errors="coerce")
    if "SN" in df.columns:
        df = df[df["SN"].astype(str).apply(clean_serial_number) == serial_number]
    if df.empty: return []
    df = df.sort_values("Sampling Date")

    parameters = [
        "Power factor @25C/60Hz (%)","IFT","Breakdown voltage @2.5mm (kV)",
        "Humidity in cellulose (%)","Water Content @20 °C","Acidity (mg KOH/g)",
        "Oil Quality Index","Color ASTM","DBDS","Indirect DP-value",
        "Hydrogen (H2)","Methane (CH4)","Acetylene (C2H2)","Ethylene (C2H4)",
        "Ethane (C2H6)","Carbon Monoxide (CO)","Carbon Dioxide (CO2)","TDCG","CO2/CO Ratio"
    ]
    figs = []
    for param in parameters:
        if param not in df.columns: continue
        y = pd.to_numeric(df[param], errors="coerce")
        dates = df["Sampling Date"]
        mask = (~dates.isna()) & (~y.isna())
        if mask.sum()<2: continue
        x_dates = dates[mask]; y_vals = y[mask].values
        x_num = mdates.date2num(x_dates.dt.to_pydatetime())
        x_smooth, y_smooth = _catmull_rom_chain(x_num, y_vals)
        x_smooth_dates = mdates.num2date(x_smooth)

        fig, ax = plt.subplots(figsize=(5,3.5), facecolor="#FAFBFF")
        ax.plot(x_smooth_dates, y_smooth, lw=2, color=TREND_LINE_COLOR, alpha=0.9)
        ax.plot(x_dates, y_vals, "o", ms=5, mfc="#FAFBFF", mec=TREND_LINE_COLOR, mew=1.3, zorder=3)

        ymin = np.nanmin(np.concatenate([y_vals,y_smooth]))
        ymax = np.nanmax(np.concatenate([y_vals,y_smooth]))
        if ymin==ymax:
            ymin_plot,ymax_plot = (ymin-1,ymax+1) if ymin==0 else (ymin*0.9,ymax*1.1)
        else:
            rng=ymax-ymin; pad=0.15*rng
            ymin_plot,ymax_plot = ymin-pad, ymax+pad
        ax.set_ylim(ymin_plot, ymax_plot)
        _add_trend_bands(ax, param, ymin_plot, ymax_plot)
        ax.set_title(param, fontsize=8, fontweight="bold")
        ax.set_xlabel("Date", fontsize=7); ax.set_ylabel(param, fontsize=7)
        ax.tick_params(axis="x", labelsize=7, rotation=30)
        ax.tick_params(axis="y", labelsize=7)
        ax.grid(True, ls="--", lw=0.5, color="#cccccc")
        fig.tight_layout()
        figs.append((param, fig))
    return figs

# ══════════════════════════════════════════════════════════════════════════════
#  PROCESS SAMPLE (core pipeline)
# ══════════════════════════════════════════════════════════════════════════════

def process_sample(extracted_data, missing_overrides=None):
    """Run the full pipeline: normalize → backfill → ML → score → save."""
    if missing_overrides:
        extracted_data.update(missing_overrides)

    _normalize_sampling_date(extracted_data)
    raw_sn = extracted_data.get("Serial Number:","")
    sn_key = clean_serial_number(raw_sn)
    extracted_data["Serial Number:"] = sn_key

    sap_key = extracted_data.get("SAP No.:","")
    curr_date_str = extracted_data.get("Sampling Date:","")

    raw_2fal_in = extracted_data.get("2-Furfural (2FAL)", None)
    try:
        raw_2fal_num = float(raw_2fal_in) if not _is_missing(raw_2fal_in) else None
    except: raw_2fal_num = None
    extracted_data["Raw 2-Furfural (2FAL)"] = raw_2fal_num

    raw_dp_in = extracted_data.get("Indirect DP-value", None)
    try:
        raw_dp_num = int(float(raw_dp_in)) if not _is_missing(raw_dp_in) else None
    except: raw_dp_num = None
    extracted_data["Raw Indirect DP"] = raw_dp_num

    backfill_metadata(extracted_data, curr_date_str)
    loc, sub = standardize_location(extracted_data.get("Location:",""), extracted_data.get("Substation:",""))
    extracted_data["Location:"] = loc; extracted_data["Substation:"] = sub

    sn_key  = extracted_data.get("Serial Number:","")
    sap_key = extracted_data.get("SAP No.:","")

    # Find missing required fields that need user input
    missing_fields = []
    for p in REQUIRED_PARAMS:
        ptype = REQ_TYPES[p]
        val = extracted_data.get(p)
        if _needs_prompt(val, ptype):
            bf = backfill_numeric(p, sap_key, sn_key, curr_date_str)
            if bf is not None and not _needs_prompt(bf, ptype):
                extracted_data[p] = bf
            else:
                missing_fields.append(p)

    if missing_fields:
        return None, missing_fields  # caller must collect and re-call

    # 2FAL correction from history
    df_hist = _load_db_cached()
    prev_raw_2fal = None
    prev_actual_2fal = None
    if df_hist is not None and "SN" in df_hist.columns:
        sub_hist = df_hist[df_hist["SN"].astype(str).apply(clean_serial_number)==sn_key].copy()
        if "Sampling From" in sub_hist.columns:
            sub_hist = sub_hist[sub_hist["Sampling From"].astype(str).str.lower().str.contains("bottom",na=False)]
        if not sub_hist.empty:
            curr_dt = pd.to_datetime(curr_date_str, dayfirst=True, errors="coerce")
            sub_hist["_dt"] = pd.to_datetime(sub_hist["Sampling Date"], errors="coerce")
            if not pd.isna(curr_dt):
                sub_hist = sub_hist[sub_hist["_dt"]<curr_dt]
            if not sub_hist.empty:
                sub_hist = sub_hist.sort_values("_dt")
                last = sub_hist.iloc[-1]
                try: prev_raw_2fal = float(last.get("Raw 2-Furfural (2FAL)",last.get("2-Furfural (2FAL)",0)))
                except: prev_raw_2fal = None
                try: prev_actual_2fal = float(last.get("2-Furfural (2FAL)",0))
                except: prev_actual_2fal = 0.0

    curr_raw_2fal = _to_float(extracted_data.get("2-Furfural (2FAL)"), 0.0)
    maint_reset = False
    if prev_raw_2fal is not None and prev_raw_2fal > 1:
        drop = (prev_raw_2fal - curr_raw_2fal)/prev_raw_2fal if prev_raw_2fal>0 else 0
        maint_reset = drop >= RAW_2FAL_DROP_FOR_MAINT
    latest_actual = (prev_actual_2fal or 0.0) if not maint_reset else 0.0
    actual_2fal = max(curr_raw_2fal + latest_actual, prev_actual_2fal or 0.0)
    extracted_data["2-Furfural (2FAL)"] = actual_2fal
    maint_note = MAINT_NOTE_TEXT if maint_reset else ""
    extracted_data["Maintenance Note"] = maint_note

    if actual_2fal > 0:
        dp_updated = int(round((1.5655-math.log10(actual_2fal))/0.0035))
    else: dp_updated = 0
    extracted_data["Indirect DP-value"] = dp_updated

    tension_str, voltage_level = parse_tension_hvlv(extracted_data.get("Tension [HV/LV] (KV):",""))
    extracted_data["Tension [HV/LV] (KV):"] = tension_str
    if voltage_level is None: voltage_level = 0

    # Age
    manu_raw = extracted_data.get("Manufacture Year:",None)
    cy = _clean_year_string(manu_raw)
    age_val = 0
    if cy:
        try:
            manu_year = int(cy)
            age_val = max(0, datetime.now().year - manu_year)
        except: pass
    if age_val == 0 and not _is_missing(sn_key):
        bf_age = backfill_numeric("Age", sap_key, sn_key, curr_date_str)
        if bf_age is not None: age_val = _to_int(bf_age)
    extracted_data["Age"] = age_val

    # ML predictions
    oil_df = pd.DataFrame({
        "Voltage level (kV)":[voltage_level], "Age":[age_val],
        "Power factor @25C/60Hz (%)":[extracted_data.get("Power Factor @25°C/60Hz")],
        "IFT":[extracted_data.get("Interfacial Tension (IFT)")],
        "Breakdown voltage @2.5mm (kV)":[extracted_data.get("Breakdown Voltage @2.5mm- AVG")],
        "Humidity in cellulose (%)":[extracted_data.get("Humidity in Cellulose")],
        "Water Content @20 °C":[extracted_data.get("Water Content @20 °C")],
        "Acidity (mg KOH/g)":[extracted_data.get("Total Acid Number")],
        "Oil Quality Index":[extracted_data.get("Oil Quality Index")],
        "Color ASTM":[extracted_data.get("Color ASTM")],
        "DBDS":[extracted_data.get("Dibenzyl Disulfide (DBDS)")],
        "Indirect DP-value":[dp_updated],
    }).reindex(columns=FEATURES_OIL).fillna(0)
    oil_rec = str(le_oil.inverse_transform(rf_oil.predict(oil_df))[0])

    co = _to_float(extracted_data.get("Carbon Monoxide (CO)"))
    co2 = _to_float(extracted_data.get("Carbon Dioxide (CO2)"))
    h2  = _to_float(extracted_data.get("Hydrogen (H2)"))
    ch4 = _to_float(extracted_data.get("Methane (CH4)"))
    c2h6= _to_float(extracted_data.get("Ethane (C2H6)"))
    c2h2= _to_float(extracted_data.get("Acetylene (C2H2)"))
    c2h4= _to_float(extracted_data.get("Ethylene (C2H4)"))
    TDCG = h2+ch4+c2h6+c2h2+c2h4+co
    co2_co_ratio = round(co2/co,2) if co!=0 else 0
    extracted_data["TDCG"] = TDCG
    extracted_data["CO2/CO Ratio"] = co2_co_ratio

    dga_df = pd.DataFrame({
        "Hydrogen (H2)":[h2],"Methane (CH4)":[ch4],"Acetylene (C2H2)":[c2h2],
        "Ethylene (C2H4)":[c2h4],"Ethane (C2H6)":[c2h6],
        "Carbon Monoxide (CO)":[co],"Carbon Dioxide (CO2)":[co2],
        "TDCG":[TDCG],"CO2/CO Ratio":[co2_co_ratio],"Indirect DP-value":[dp_updated],
    })
    dga1_rec = str(le_dga1.inverse_transform(rf_dga1.predict(dga_df.reindex(columns=FEATURES_DGA1).fillna(0)))[0])
    dga2_rec = str(le_dga2.inverse_transform(rf_dga2.predict(dga_df.reindex(columns=FEATURES_DGA2).fillna(0)))[0])

    score_row = {
        "Breakdown voltage @2.5mm (kV)": extracted_data.get("Breakdown Voltage @2.5mm- AVG",0),
        "IFT": extracted_data.get("Interfacial Tension (IFT)",0),
        "Acidity (mg KOH/g)": extracted_data.get("Total Acid Number",0),
        "Power factor @25C/60Hz (%)": extracted_data.get("Power Factor @25°C/60Hz",0),
        "Water Content @20 °C": extracted_data.get("Water Content @20 °C",0),
        "Humidity in cellulose (%)": extracted_data.get("Humidity in Cellulose",0),
        "DBDS": extracted_data.get("Dibenzyl Disulfide (DBDS)",0),
        "Color ASTM": extracted_data.get("Color ASTM",0),
        "Oil Quality Index": extracted_data.get("Oil Quality Index",0),
        "Hydrogen (H2)": h2,"Methane (CH4)": ch4,"Acetylene (C2H2)": c2h2,
        "Ethylene (C2H4)": c2h4,"Ethane (C2H6)": c2h6,
        "Carbon Monoxide (CO)": co,"TDCG": TDCG,
        "Carbon Dioxide (CO2)": co2,"CO2/CO Ratio": co2_co_ratio,
        "Indirect DP-value": dp_updated,
    }
    final_score = score_v2(score_row)
    extracted_data["Score"] = final_score
    condition = "Good" if final_score>=6.67 else "Fair" if final_score>=3.34 else "Poor"
    extracted_data["Condition"] = condition
    extracted_data["Escalated to TCA L2"] = "Yes" if final_score<6.67 else "No"

    saved = save_to_excel(extracted_data, oil_rec, dga1_rec, dga2_rec, maint_note)
    try: recompute_transformer_history(sn_key)
    except: pass
    _load_db_cached.clear()

    return {
        "extracted": extracted_data,
        "oil_rec": oil_rec, "dga1_rec": dga1_rec, "dga2_rec": dga2_rec,
        "maint_note": maint_note, "score": final_score, "condition": condition,
        "saved": saved,
    }, []

# ══════════════════════════════════════════════════════════════════════════════
#  PDF REPORT EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def _fmt_date(v):
    try:
        if isinstance(v, pd.Timestamp) and not pd.isna(v): return v.strftime("%d/%m/%Y")
    except: pass
    return str(v) if not _is_missing(v) else "NA"

def export_pdf_report(sn):
    if not DB_PATH.is_file(): return None, "Database not found"
    df = pd.read_excel(DB_PATH)
    df.columns = df.columns.astype(str).str.strip()
    if "SN" not in df.columns: return None, "SN column missing"
    df["SN_clean"] = df["SN"].astype(str).apply(clean_serial_number)
    sn_clean = clean_serial_number(sn)
    df_sn = df[df["SN_clean"]==sn_clean].copy()
    if df_sn.empty: return None, f"No records for SN: {sn}"
    if "Sampling Date" in df_sn.columns:
        df_sn["Sampling Date_dt"] = pd.to_datetime(df_sn["Sampling Date"], dayfirst=True, errors="coerce")
    else:
        df_sn["Sampling Date_dt"] = pd.NaT
    df_sn = df_sn.sort_values("Sampling Date_dt")
    latest = df_sn.iloc[-1]
    styles = getSampleStyleSheet()
    hdr_style = ParagraphStyle("hdr", parent=styles["BodyText"], fontSize=7, leading=9)
    rec_style = ParagraphStyle("rec", parent=styles["BodyText"], fontSize=7, leading=9)

    info_lines = [
        f"<b>SN:</b> {sn}",
        f"<b>SAP:</b> {latest.get('SAP No','')}",
        f"<b>Location:</b> {latest.get('Location','')}",
        f"<b>Substation:</b> {latest.get('Substation','')}",
        f"<b>Tag Number:</b> {latest.get('Tag Number','')}",
        f"<b>Manufacture Name:</b> {latest.get('Manufacture Name','')}",
    ]
    key_cols = [
        ("Date","Sampling Date"),("Sampling From","Sampling From"),("Score","Score"),
        ("Condition","Condition"),("DP","Indirect DP-value"),("2FAL","2-Furfural (2FAL)"),
        ("BDV","Breakdown voltage @2.5mm (kV)"),("IFT","IFT"),("TAN","Acidity (mg KOH/g)"),
        ("PF(%)","Power factor @25C/60Hz (%)"),
    ]
    rows = []
    for _, r in df_sn.iterrows():
        rv = []
        for label, col in key_cols:
            if col=="Sampling Date": rv.append(_fmt_date(r.get("Sampling Date_dt",r.get("Sampling Date","NA"))))
            else:
                v = r.get(col,"NA")
                rv.append("NA" if _is_missing(v) else str(v))
        rows.append(rv)
        rec_text = (f"Oil: {r.get('Oil Quality Recommendations','')} | "
                    f"DGA1: {r.get('Group1 DGA Recommendations','')} | "
                    f"DGA2: {r.get('Group2 DGA Recommendations','')}")
        maint = _dedupe_pipe_text(str(r.get("Maintenance Note","") or "").strip())
        if maint: rec_text += f" | Maintenance: {maint}"
        rows.append([Paragraph(rec_text, rec_style)]+[""]*(len(key_cols)-1))

    figs = generate_trend_graphs(sn_clean)
    def _fig_to_img(fig, width=520):
        bio = io.BytesIO()
        fig.savefig(bio, format="png", dpi=150, bbox_inches="tight")
        bio.seek(0)
        img = RLImage(bio)
        img.drawWidth = width
        aspect = img.imageHeight/float(img.imageWidth) if img.imageWidth else 0.75
        img.drawHeight = width*aspect
        return img

    bio_out = io.BytesIO()
    doc = SimpleDocTemplate(bio_out, pagesize=A4, rightMargin=14, leftMargin=14, topMargin=20, bottomMargin=20)
    story = []
    story.append(Paragraph(f"Transformer Report – SN: {sn}", styles["Title"]))
    story.append(Spacer(1,10))
    story.append(Paragraph("<b>Transformer Information</b>", styles["Heading2"]))
    for line in info_lines: story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1,10))
    story.append(Paragraph("<b>Key Parameters (All Samples)</b>", styles["Heading2"]))
    header_row = [Paragraph(lbl,hdr_style) for lbl,_ in key_cols]
    table_data = [header_row]+rows
    avail_w = doc.width; ncols = len(header_row)
    first = avail_w*0.18; rest = (avail_w-first)/(ncols-1)
    col_widths = [first]+[rest]*(ncols-1)
    tbl = Table(table_data, repeatRows=1, colWidths=col_widths)
    ts = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),rl_colors.lightgrey),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7),("GRID",(0,0),(-1,-1),0.25,rl_colors.grey),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(0,0),(-1,-1),"CENTER"),
    ])
    for i in range(2,len(table_data),2):
        ts.add("BACKGROUND",(0,i),(-1,i),rl_colors.lightgrey)
        ts.add("SPAN",(0,i),(-1,i)); ts.add("ALIGN",(0,i),(-1,i),"LEFT")
        ts.add("FONTNAME",(0,i),(-1,i),"Helvetica-Bold")
    tbl.setStyle(ts)
    story.append(tbl); story.append(PageBreak())
    story.append(Paragraph("<b>Trends</b>", styles["Heading2"])); story.append(Spacer(1,10))
    img_w = (doc.width-12)/2
    trend_imgs = [_fig_to_img(fig, width=img_w) for _, fig in figs]
    rows_grid = []
    for i in range(0, len(trend_imgs), 2):
        row = trend_imgs[i:i+2]
        if len(row)<2: row += [""]*(2-len(row))
        rows_grid.append(row)
    if rows_grid:
        tt = Table(rows_grid, colWidths=[img_w,img_w])
        tt.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
        story.append(tt)
    doc.build(story)
    bio_out.seek(0)
    return bio_out.getvalue(), None

# ══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD CHARTS
# ══════════════════════════════════════════════════════════════════════════════

def _latest_per_sn(df):
    if df is None or df.empty or "SN" not in df.columns or "Sampling Date" not in df.columns: return pd.DataFrame()
    return df.sort_values("Sampling Date").groupby("SN",as_index=False).tail(1)

def _ensure_condition(row):
    cond = row.get("Condition","")
    if isinstance(cond,str) and cond.strip(): return cond.strip()
    try:
        s = float(row.get("Score",0))
        return "Good" if s>=6.67 else "Fair" if s>=3.34 else "Poor"
    except: return ""

def render_dashboard(df, date_from=None, date_to=None):
    if df is None or df.empty:
        st.info("No data in Database.xlsx yet.")
        return
    df = df.copy()
    if "Sampling Date" in df.columns:
        df["Sampling Date"] = pd.to_datetime(df["Sampling Date"], errors="coerce")
        if date_from: df = df[df["Sampling Date"]>=pd.to_datetime(date_from, dayfirst=True, errors="coerce")]
        if date_to:   df = df[df["Sampling Date"]<=pd.to_datetime(date_to,   dayfirst=True, errors="coerce")]
    if df.empty: st.info("No data matches the selected date range."); return

    latest = _latest_per_sn(df)
    latest["_Condition"] = latest.apply(_ensure_condition, axis=1)
    n_samples = len(df); n_trans = df["SN"].nunique() if "SN" in df.columns else 0
    n_good = (latest["_Condition"]=="Good").sum()
    n_fair = (latest["_Condition"]=="Fair").sum()
    n_poor = (latest["_Condition"]=="Poor").sum()

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("📄 Total Samples", n_samples)
    c2.metric("⚡ Transformers", n_trans)
    c3.metric("✅ Good", n_good)
    c4.metric("⚠️ Fair", n_fair)
    c5.metric("🚨 Poor", n_poor)
    st.divider()

    base_colors = ["#42A5F5","#26A69A","#7E57C2","#FFB74D","#26C6DA","#8D6E63"]
    cols = st.columns(2)

    with cols[0]:
        if "Location" in df.columns:
            loc_counts = df.assign(Location=df["Location"].fillna("Unknown").astype(str)).groupby("Location").size().sort_index()
            n_loc = len(loc_counts)
            fig,ax = plt.subplots(figsize=(5,3.5),facecolor="#FAFBFF")
            clrs = (base_colors*((n_loc//len(base_colors))+1))[:n_loc]
            bars = ax.bar(loc_counts.index, loc_counts.values, color=clrs, edgecolor="#263238", lw=0.9)
            ax.set_title("Transformers per Location",fontsize=10,fontweight="bold")
            ax.set_xlabel("Location",fontsize=9); ax.set_ylabel("Count",fontsize=9)
            ax.tick_params(axis="x",rotation=35,labelsize=8); ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            try: ax.bar_label(bars,padding=3,fontsize=9)
            except: pass
            ax.grid(True,axis="y",ls="--",lw=0.4,alpha=0.4); fig.tight_layout()
            st.pyplot(fig); plt.close(fig)

    with cols[1]:
        cond_counts = latest["_Condition"].value_counts()
        fig,ax = plt.subplots(figsize=(4,3.5),facecolor="#FAFBFF")
        pie_colors = {"Good":"#4CAF50","Fair":"#FFA726","Poor":"#EF5350"}
        labels = cond_counts.index.tolist()
        pie_c = [pie_colors.get(l,"#90A4AE") for l in labels]
        wedges,texts,autotexts = ax.pie(cond_counts.values,labels=labels,colors=pie_c,
            autopct="%1.0f%%",startangle=90,wedgeprops=dict(edgecolor="white",linewidth=2))
        ax.set_title("Condition Distribution (Latest Sample)",fontsize=10,fontweight="bold")
        fig.tight_layout()
        st.pyplot(fig); plt.close(fig)

    if "Sampling Date" in df.columns and not df["Sampling Date"].isna().all():
        fig,ax = plt.subplots(figsize=(10,3),facecolor="#FAFBFF")
        df["_year_month"] = df["Sampling Date"].dt.to_period("M").astype(str)
        monthly = df.groupby("_year_month").size()
        ax.bar(monthly.index, monthly.values, color="#42A5F5", edgecolor="#1565C0", lw=0.7)
        ax.set_title("Samples per Month",fontsize=10,fontweight="bold")
        ax.set_xlabel("Month"); ax.set_ylabel("Count")
        ax.tick_params(axis="x",rotation=45,labelsize=7)
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.grid(True,axis="y",ls="--",lw=0.4,alpha=0.4); fig.tight_layout()
        st.pyplot(fig); plt.close(fig)

def _is_alarming(param, value):
    if param not in TREND_LIMITS or value is None: return False
    try: v=float(value)
    except: return False
    low, high = TREND_LIMITS[param]["alarming"]
    if low is None and high is None: return False
    if low is None:  return v<high
    if high is None: return v>low
    return low<=v<=high

def get_need_attention_rows(selected_params=None):
    df = _load_db_cached()
    if df is None or df.empty: return []
    if selected_params is None: selected_params = list(TREND_LIMITS.keys())
    latest = _latest_per_sn(df)
    if latest.empty: return []
    rows = []
    for _, row in latest.iterrows():
        pieces = []
        for col in selected_params:
            if col not in row.index: continue
            val = row[col]
            if pd.isna(val): continue
            if not _is_alarming(col, val): continue
            lbl = PARAM_LABELS.get(col, col)
            unit = PARAM_UNITS.get(col, "")
            try: vf = float(val); vs = str(int(vf)) if abs(vf-int(vf))<1e-6 else f"{vf:.2f}"
            except: vs = str(val)
            pieces.append(f"{lbl} ({vs} {unit})".strip())
        if not pieces: continue
        dt = row.get("Sampling Date", pd.NaT)
        ds = dt.strftime("%d/%m/%Y") if isinstance(dt,pd.Timestamp) and not pd.isna(dt) else ""
        rows.append({"SN":row.get("SN",""),"SAP":row.get("SAP No",""),
                     "Sample No":row.get("Sample No",""),
                     "Sampling From":row.get("Sampling From",""),
                     "Sampling Date":ds,"Alarms":", ".join(pieces)})
    return rows

# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
  .main .block-container { padding-top: 1rem; }
  .stMetric { background:#f7f9fe; border-radius:8px; padding:10px; }
  .stTabs [data-baseweb="tab"] { font-size:15px; font-weight:600; }
</style>
""", unsafe_allow_html=True)

st.title("⚡ TOSA – Transformer Oil Sample Analyzer")

tab_upload, tab_sample, tab_dashboard, tab_attention, tab_trends, tab_db = st.tabs([
    "📤 Upload Sample", "📋 Sample Results", "📊 Dashboard",
    "🚨 Need Attention", "📈 Trends", "🗄️ Database"
])

# ── SESSION STATE ─────────────────────────────────────────────────────────────
for key in ["result","extracted","missing_fields","pending_overrides","last_sn"]:
    if key not in st.session_state:
        st.session_state[key] = None if key != "missing_fields" else []

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1 – UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
with tab_upload:
    st.subheader("Upload Lab Report")
    st.write("Upload a **PDF** or **Excel** lab report. The app extracts parameters, runs ML models, scores the transformer, and saves to the shared database.")

    uploaded = st.file_uploader("Choose file", type=["pdf","xlsx","xls"], label_visibility="collapsed")

    if uploaded:
        file_bytes = uploaded.read()
        file_name = uploaded.name.lower()
        extracted = {}

        with st.spinner("Extracting parameters…"):
            try:
                if file_name.endswith(".pdf"):
                    doc_pdf = fitz.open(stream=file_bytes, filetype="pdf")
                    text = "".join(page.get_text() for page in doc_pdf)
                    extracted = extract_parameters(text)
                else:
                    df_xl = pd.read_excel(io.BytesIO(file_bytes))
                    df_xl.columns = df_xl.columns.astype(str).str.strip()
                    extracted = extract_from_excel_row(df_xl.iloc[0].to_dict())
            except Exception as e:
                st.error(f"Failed to read file: {e}")
                st.stop()

        st.session_state.extracted = extracted
        st.session_state.missing_fields = []
        st.session_state.pending_overrides = {}
        st.session_state.result = None

        st.success("File read successfully. Reviewing extracted data…")
        with st.expander("📄 Extracted Parameters", expanded=True):
            col1, col2 = st.columns(2)
            items = list(extracted.items())
            half = len(items)//2
            with col1:
                for k,v in items[:half]: st.write(f"**{k}** → `{v}`")
            with col2:
                for k,v in items[half:]: st.write(f"**{k}** → `{v}`")

        # Run pipeline (first pass)
        result, missing = process_sample(dict(extracted))
        st.session_state.missing_fields = missing

        if not missing:
            st.session_state.result = result
            if result["saved"] is False:
                st.warning("⚠️ This sample already exists in the database (duplicate). Results shown but not saved again.")
            else:
                st.success("✅ Sample processed and saved to database!")
        else:
            st.warning(f"⚠️ {len(missing)} parameter(s) could not be extracted or backfilled. Please enter them below.")

    # Missing parameters form
    if st.session_state.missing_fields and st.session_state.result is None:
        st.subheader("📝 Enter Missing Parameters")
        sn  = st.session_state.extracted.get("Serial Number:","NA")
        sap = st.session_state.extracted.get("SAP No.:","NA")
        dat = st.session_state.extracted.get("Sampling Date:","NA")
        st.info(f"Asset: **{sn}** | SAP: **{sap}** | Date: **{dat}**")

        with st.form("missing_params_form"):
            overrides = {}
            for p in st.session_state.missing_fields:
                ptype = REQ_TYPES.get(p,"float")
                label = f"{p} ({ptype})"
                existing = st.session_state.pending_overrides.get(p,"")
                val_str = st.text_input(label, value=str(existing) if existing else "")
                overrides[p] = val_str

            if not st.session_state.extracted.get("Sampling Date:") or st.session_state.extracted.get("Sampling Date:") == "NA":
                date_input = st.text_input("Sampling Date (dd/mm/yyyy)")
                overrides["Sampling Date:"] = date_input

            submitted = st.form_submit_button("Submit & Process")

        if submitted:
            parsed = {}
            errors = []
            for p, val_str in overrides.items():
                if p == "Sampling Date:":
                    if val_str.strip():
                        try: parsed[p] = _parse_date_strict(val_str)
                        except Exception as e: errors.append(f"Date: {e}")
                    continue
                ptype = REQ_TYPES.get(p,"float")
                try:
                    if ptype=="int": parsed[p] = int(float(val_str))
                    else:
                        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", val_str)
                        if not m: raise ValueError("No number found")
                        parsed[p] = float(m.group(0))
                except Exception as e:
                    errors.append(f"'{p}': {e}")
            if errors:
                for e in errors: st.error(e)
            else:
                result, missing2 = process_sample(dict(st.session_state.extracted), parsed)
                if missing2:
                    st.error(f"Still missing: {missing2}")
                    st.session_state.missing_fields = missing2
                else:
                    st.session_state.result = result
                    st.session_state.missing_fields = []
                    if result["saved"] is False:
                        st.warning("Duplicate sample – results shown but not saved again.")
                    else:
                        st.success("✅ Sample processed and saved!")

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2 – SAMPLE RESULTS
# ══════════════════════════════════════════════════════════════════════════════
with tab_sample:
    result = st.session_state.result
    if result is None:
        st.info("Upload and process a sample to see results here.")
    else:
        ed = result["extracted"]
        sn = ed.get("Serial Number:","NA")
        score = result["score"]
        condition = result["condition"]
        st.session_state.last_sn = clean_serial_number(sn)

        color_map = {"Good":"🟢","Fair":"🟡","Poor":"🔴"}
        icon = color_map.get(condition,"⚪")
        st.subheader(f"{icon} Transformer: {sn} — {condition}  (Score: {score})")

        col1, col2, col3 = st.columns(3)
        col1.metric("Health Score", score)
        col2.metric("Condition", condition)
        col3.metric("Escalated", ed.get("Escalated to TCA L2","No"))

        st.divider()
        st.subheader("🔬 ML Recommendations")
        rc1,rc2,rc3 = st.columns(3)
        rc1.info(f"**Oil Quality**\n\n{result['oil_rec']}")
        rc2.info(f"**DGA Group 1**\n\n{result['dga1_rec']}")
        rc3.info(f"**DGA Group 2**\n\n{result['dga2_rec']}")
        if result["maint_note"]:
            st.warning(f"⚙️ {result['maint_note']}")

        st.divider()
        st.subheader("📋 Extracted Parameters")
        display = {
            "Serial Number": ed.get("Serial Number:"), "SAP No.": ed.get("SAP No.:"),
            "Location": ed.get("Location:"), "Substation": ed.get("Substation:"),
            "Sampling Date": ed.get("Sampling Date:"), "Sampling From": ed.get("Sampling From:"),
            "Tension [HV/LV]": ed.get("Tension [HV/LV] (KV):"), "Age (years)": ed.get("Age"),
            "Power Factor (%)": ed.get("Power Factor @25°C/60Hz"),
            "IFT": ed.get("Interfacial Tension (IFT)"),
            "BDV (kV)": ed.get("Breakdown Voltage @2.5mm- AVG"),
            "Humidity in Cellulose (%)": ed.get("Humidity in Cellulose"),
            "Water Content (ppm)": ed.get("Water Content @20 °C"),
            "Total Acid Number": ed.get("Total Acid Number"),
            "Oil Quality Index": ed.get("Oil Quality Index"),
            "Color ASTM": ed.get("Color ASTM"),
            "DBDS (ppm)": ed.get("Dibenzyl Disulfide (DBDS)"),
            "2-FAL (ppm)": ed.get("2-Furfural (2FAL)"),
            "DP-value": ed.get("Indirect DP-value"),
            "H₂ (ppm)": ed.get("Hydrogen (H2)"),
            "CH₄ (ppm)": ed.get("Methane (CH4)"),
            "C₂H₂ (ppm)": ed.get("Acetylene (C2H2)"),
            "C₂H₄ (ppm)": ed.get("Ethylene (C2H4)"),
            "C₂H₆ (ppm)": ed.get("Ethane (C2H6)"),
            "CO (ppm)": ed.get("Carbon Monoxide (CO)"),
            "CO₂ (ppm)": ed.get("Carbon Dioxide (CO2)"),
            "TDCG (ppm)": ed.get("TDCG"),
            "CO₂/CO Ratio": ed.get("CO2/CO Ratio"),
        }
        df_disp = pd.DataFrame(display.items(), columns=["Parameter","Value"])
        st.dataframe(df_disp, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("📑 Export PDF Report")
        if st.button("Generate PDF Report"):
            with st.spinner("Building report…"):
                pdf_bytes, err = export_pdf_report(sn)
            if err: st.error(err)
            else:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    "⬇️ Download PDF",
                    data=pdf_bytes,
                    file_name=f"Transformer_Report_{clean_serial_number(sn)}_{ts}.pdf",
                    mime="application/pdf"
                )

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 3 – DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
with tab_dashboard:
    st.subheader("📊 Fleet Dashboard")
    df_db = _load_db_cached()
    col_f1, col_f2, col_f3 = st.columns([2,2,1])
    with col_f1: date_from = st.text_input("From (dd/mm/yyyy)", key="dash_from")
    with col_f2: date_to   = st.text_input("To (dd/mm/yyyy)",   key="dash_to")
    with col_f3:
        st.write(""); st.write("")
        if st.button("🔄 Refresh Dashboard"): _load_db_cached.clear(); st.rerun()
    render_dashboard(df_db, date_from or None, date_to or None)

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 4 – NEED ATTENTION
# ══════════════════════════════════════════════════════════════════════════════
with tab_attention:
    st.subheader("🚨 Transformers Needing Attention")
    st.write("Shows the latest sample per transformer where at least one parameter is in the **alarming** zone.")

    with st.expander("⚙️ Select parameters to monitor"):
        all_params = list(TREND_LIMITS.keys())
        selected = st.multiselect("Monitored parameters", all_params, default=all_params, key="attention_params")

    if st.button("🔄 Refresh", key="refresh_attention"): _load_db_cached.clear()
    rows = get_need_attention_rows(selected)
    if not rows:
        st.success("✅ No transformers with alarming values for the selected parameters.")
    else:
        df_att = pd.DataFrame(rows)
        st.dataframe(df_att, use_container_width=True, hide_index=True)
        st.caption(f"{len(rows)} transformer(s) with alarming values.")

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 5 – TRENDS
# ══════════════════════════════════════════════════════════════════════════════
with tab_trends:
    st.subheader("📈 Transformer Trend Analysis")
    df_db = _load_db_cached()
    if df_db is None or df_db.empty:
        st.info("No database found. Process a sample first.")
    else:
        all_sns = sorted(df_db["SN"].dropna().astype(str).unique().tolist()) if "SN" in df_db.columns else []
        default_sn = st.session_state.last_sn if st.session_state.last_sn in all_sns else (all_sns[0] if all_sns else "")
        selected_sn = st.selectbox("Select Transformer (SN)", all_sns, index=all_sns.index(default_sn) if default_sn in all_sns else 0)

        if selected_sn:
            with st.spinner("Generating trend charts…"):
                figs = generate_trend_graphs(selected_sn, db_df=df_db)
            if not figs:
                st.info("Not enough data for trends (need ≥ 2 bottom-tank samples for this SN).")
            else:
                DGA_PARAMS = {"Hydrogen (H2)","Methane (CH4)","Acetylene (C2H2)","Ethylene (C2H4)",
                              "Ethane (C2H6)","Carbon Monoxide (CO)","Carbon Dioxide (CO2)","TDCG","CO2/CO Ratio"}
                oil_figs = [(p,f) for p,f in figs if p not in DGA_PARAMS]
                dga_figs = [(p,f) for p,f in figs if p in DGA_PARAMS]

                if oil_figs:
                    st.markdown("#### 🛢️ Oil Quality Parameters")
                    cols_per_row = 3
                    for i in range(0, len(oil_figs), cols_per_row):
                        cols = st.columns(cols_per_row)
                        for j, (param, fig) in enumerate(oil_figs[i:i+cols_per_row]):
                            with cols[j]: st.pyplot(fig)
                        for fig in [f for _,f in oil_figs[i:i+cols_per_row]]: plt.close(fig)
                if dga_figs:
                    st.markdown("#### ⚗️ DGA Parameters")
                    for i in range(0, len(dga_figs), cols_per_row):
                        cols = st.columns(cols_per_row)
                        for j, (param, fig) in enumerate(dga_figs[i:i+cols_per_row]):
                            with cols[j]: st.pyplot(fig)
                        for fig in [f for _,f in dga_figs[i:i+cols_per_row]]: plt.close(fig)

                st.divider()
                st.markdown("#### 📑 Export PDF Report")
                if st.button("Generate PDF Report", key="trends_pdf"):
                    with st.spinner("Building report…"):
                        pdf_bytes, err = export_pdf_report(selected_sn)
                    if err: st.error(err)
                    else:
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        st.download_button(
                            "⬇️ Download PDF",
                            data=pdf_bytes,
                            file_name=f"Transformer_Report_{selected_sn}_{ts}.pdf",
                            mime="application/pdf"
                        )

# ══════════════════════════════════════════════════════════════════════════════
#  TAB 6 – DATABASE
# ══════════════════════════════════════════════════════════════════════════════
with tab_db:
    st.subheader("🗄️ Full Database")
    col_r1, col_r2 = st.columns([1,5])
    with col_r1:
        if st.button("🔄 Refresh"): _load_db_cached.clear(); st.rerun()

    df_db = _load_db_cached()
    if df_db is None or df_db.empty:
        st.info("No database yet. Process a sample to create it.")
    else:
        # Filters
        sn_filter = st.text_input("Filter by SN (partial match)", key="db_sn_filter")
        if sn_filter:
            df_db = df_db[df_db["SN"].astype(str).str.contains(sn_filter, case=False, na=False)]

        st.dataframe(df_db, use_container_width=True, height=500)
        st.caption(f"{len(df_db)} rows")

        # Download
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df_db.to_excel(w, index=False)
        buf.seek(0)
        st.download_button(
            "⬇️ Download Database (Excel)",
            data=buf.getvalue(),
            file_name="Database.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Upload updated database
        st.divider()
        st.markdown("**⬆️ Upload Updated Database**")
        new_db = st.file_uploader("Upload Database.xlsx to replace current", type=["xlsx"], key="db_upload")
        if new_db:
            if st.button("✅ Confirm Replace Database"):
                DB_PATH.write_bytes(new_db.read())
                _load_db_cached.clear()
                st.success("Database replaced successfully!")
                st.rerun()
